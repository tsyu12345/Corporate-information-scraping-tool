from os import replace
from openpyxl.worksheet.dimensions import SheetDimension
import requests as rq
from selenium import webdriver
from selenium.common.exceptions import InvalidArgumentException, NoSuchElementException, TimeoutException
from openpyxl import styles as pxstyle
import openpyxl as px 
from bs4 import BeautifulSoup as bs 
import time
import re
from selenium.webdriver.common.keys import Keys

class Job:

    def __init__(self, driverPath):
        options = webdriver.ChromeOptions()     
        #options.add_argument('--headless')
        #options.add_argument('--no-sandbox')
        #options.add_argument('--disable-gpu')        
        self.driver = webdriver.Chrome(executable_path=driverPath, options=options)
        #self.driver.set_window_size(1128, 768)
        
            
    def scrap(self, url):
        try:
            self.driver.get(url)
            time.sleep(2)
        except:
            return False
        Link_names = [
            "会社案内", 
            "会社概要", 
            "会社紹介",
            "会社情報", 
            "about", 
            "ABOUT", 
            "about us", 
            "ABOUT US", 
            "About",
            "私たちについて", 
            "会社について", 
            "店舗案内", 
            "店舗紹介", 
            "当店について", 
            "院案内", 
            "院について", 
            "院紹介", 
            "院概要",
            "shop",
            "Shop",
            "SHOP",
            "shop info",
            "Shop Info",
            "INFOMATION",
            "infomation",
            "教室案内",
            "スクール案内",
            "教室について",
            "スクールについて",
            "教室概要",
            "施設紹介",
            "施設概要",
            "施設案内",
        ]
        
        for i in Link_names:
            try:
                try:
                    element = self.driver.find_element_by_link_text(i)
                    element.click()
                    break
                except:
                    pass
            except NoSuchElementException:
                pass
        time.sleep(2)
        html = self.driver.page_source
        soup = bs(html, 'lxml')
        try:
            table = soup.find('tbody')
            print(table)
        except:
            print("テーブルが見つかりません")
            return False
        try:
            all_text = table.get_text()
            all_text_list=all_text.split("\n")
            for i, text in enumerate(all_text_list):
                if text is "":
                    all_text_list.pop(i)
            com_name_list = ["会社名", "社名", "商号", "屋号", "店舗名", "店名", "名称", "教室名", "院名", "施設名"]
            com_tel_list = ["電話番号", "TEL", "tel", "電話", "連絡先", "Tel"]
            com_add_list = ["所在地", "住所", "本社所在地", "本社住所"]
            print(all_text_list)
            com_name = self.load_info(com_name_list, all_text_list)
            com_tel = self.load_info(com_tel_list, all_text_list)
            com_add = self.load_info(com_add_list, all_text_list)
            print(com_name, com_tel, com_add)
            return com_name, com_tel, com_add
        except:
            return False
    
    def load_info(self, menu_list, all_text_data):
        for i, text in enumerate(all_text_data):
            for j in menu_list:
                if j in text:
                    return all_text_data[i+1]

    def write_excel(self, index, datas):
        try:
            file_name = "./【高林様】抽出依頼リスト - コピー.xlsx"
            book = px.load_workbook(file_name)
            sheet = book.worksheets[0]
            #社名入力
            sheet["D" + str(index)].value = datas[0]
            #tel入力
            sheet["E" + str(index)].value = datas[1]
            #住所入力
            all_address = datas[2]
            all_address = re.sub('〒[0-9]{3}-[0-9]{4}', "", all_address)
            pre_name = re.search('東京都|北海道|(?:京都|大阪)府|.{2,3}県' , all_address)
            address_com = re.split('東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)#県名とそれ以降を分離
            add1 = pre_name.group()#県
            add2 = address_com[1]#それ以降
            sheet["F" + str(index)].value = add1
            sheet["G" + str(index)].value = add2
            book.save(file_name)
            print("\t==書き込み情報==")
            print("\t" + sheet["D" + str(index)].value)
            print("\t" + sheet["E" + str(index)].value)
            print("\t" + sheet["F" + str(index)].value)
            print("\t" + sheet["G" + str(index)].value)
        except:
            print("writing error!!")
            return False

    def check(self, index):#形式不備の自動チェック
            file_name = "./【高林様】抽出依頼リスト - コピー.xlsx"
            book = px.load_workbook(file_name)
            fill_yel = pxstyle.PatternFill(patternType='solid', fgColor='FFFF00', bgColor='FFFF00')
            sheet = book.worksheets[0]
            #社名
            if sheet["D" + str(index)].value == None:
                sheet["D" + str(index)].fill = fill_yel
            #TEL
            try:
                match = re.match('^0\d{2,3}-\d{1,4}-\d{4}$', sheet["E" + str(index)].value)
            except:
                match = False
            if match:
                print("No." + str(index) + "TEL number is OK")
            else:
                sheet["E" + str(index)].fill = fill_yel
            #住所（県名）
            if sheet["F" + str(index)].value == None:
                sheet["F" + str(index)].fill = fill_yel
            try:
                if " " in sheet["F" + str(index)].value:
                    sheet["F" + str(index)].value = replace(" ", "")
                if "　" in sheet["F" + str(index)].value:
                    sheet["F" + str(index)].value = replace("　", "")
                match = re.match('東京都|北海道|(?:京都|大阪)府|.{2,3}県', sheet["F" + str(index)].value)
                if match:
                    print("No." + str(index) + "prefecture name is OK")
                else:
                    sheet["F" + str(index)].fill = fill_yel
            except:
                pass
            #住所の数字を半角へ変換
            if sheet["F" + str(index)].value == None:
                sheet["F" + str(index)].fill = fill_yel
            try:
                if " " in sheet["F" + str(index)].value:
                    sheet["G" + str(index)].value = replace(" ", "")
                if "　" in sheet["F" + str(index)].value:
                    sheet["G" + str(index)].value = replace("　", "")
                trans_table = str.maketrans({"１":"1", "２":"2", "３":"3", "４":"4", "５":"5", "６":"6", "７":"7", "８":"8", "９":"9"})
                text = sheet["G" + str(index)].value
                text.translate(trans_table)
            except:
                pass
            book.save(file_name)
    
job = Job('chromedriver_win32/chromedriver.exe')
def main(index):
    file_name = "./【高林様】抽出依頼リスト - コピー.xlsx"
    book = px.load_workbook(file_name)
    sheet = book.worksheets[0]
    if sheet["D" + str(index)].value == None:
        print("loading No." + str(index)) 
        url = "http://www." + sheet["C" + str(index)].value
        datas = job.scrap(url)
        if datas == False:
            print("Failed")
            return False
        job.write_excel(index, datas)
    job.check(index)

for i in range(3, 2391 + 1):
    main(i)

        
 


            
