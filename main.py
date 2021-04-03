import requests as rq
from selenium import webdriver
from selenium.common.exceptions import InvalidArgumentException, NoSuchElementException, TimeoutException
from openpyxl import styles as pxstyle
import openpyxl as px 
from bs4 import BeautifulSoup as bs 
import time
import re
from selenium.webdriver.common.keys import Keys



#sheetの読み込み
file_name = "./【高林様】抽出依頼リスト - コピー.xlsx"
book = px.load_workbook(file_name)
sheet = book.worksheets[0]

"""
#サイトの存在確認
fill_red = pxstyle.PatternFill(patternType='solid', fgColor='FF0000', bgColor='FF0000')#応答がないサイトには赤色で塗る
fill_whi = pxstyle.PatternFill(patternType='solid', fgColor='ffffff', bgColor='ffffff')
for i in range(3, sheet.max_row):
    index = "C"+str(i)
    url = sheet[index].value
    print(index + " . " + url + " : ", end="")
    try:
        respons = rq.get("http://www." + url)
        if respons.status_code != 200:#閲覧できないとき
            print(respons.status_code)
            sheet[index].fill = fill_red
            sheet["D" + str(i)].value = "取得不可"
        else:#通常通り
            page_url = respons.url
            print(respons.status_code)
            print(page_url)
            sheet[index].fill = fill_whi
            #SSLかどうか（httpsで始まっているかどうか）
            if page_url.startswith("https://"):
                print("SSL 有")
                sheet["H" + str(i)].value = "有"
            else:
                sheet["H" + str(i)].value = "無"
            
    except:#そもそもサイトが存在しない？？
        print(respons.status_code)
        sheet[index].fill = fill_red
        sheet["D"+str(i)].value = "不明なエラー"
        pass
book.save(file_name)
"""


"""
#会社名がスクレイピング出来る場合にはスクレイピングで取得
driver = webdriver.Chrome(executable_path='./chromedriver_win32/chromedriver.exe')
driver.get("https://www.google.com/webhp?hl=ja&sa=X&ved=0ahUKEwiAxNDb797vAhWp3mEKHZyWAtYQPAgI")
time.sleep(1)
def com_name(domein, index):
    try:
        search = driver.find_element_by_css_selector('body > div.L3eUgb > div.o3j99.ikrT4e.om7nvf > form > div:nth-child(1) > div.A8SBwf > div.RNNXgb > div > div.a4bIc > input')
        search.send_keys(domein)
        search = driver.find_element_by_css_selector('body > div.L3eUgb > div.o3j99.ikrT4e.om7nvf > form > div:nth-child(1) > div.A8SBwf > div.FPdoLc.tfB0Bf > center > input.gNO89b')
        search.click()
    except NoSuchElementException:
        search = driver.find_element_by_css_selector('#tsf > div:nth-child(1) > div.A8SBwf > div.RNNXgb > div > div.a4bIc > input')
        search.send_keys(domein)
        search = driver.find_element_by_css_selector('#tsf > div:nth-child(1) > div.A8SBwf > div.RNNXgb > button')
        search.click()
    time.sleep(2)
    sorce = driver.page_source
    soup = bs(sorce, 'lxml')
    name = soup.select_one('#rhs > div > div.kp-blk.knowledge-panel.Wnoohf.OJXvsb > div > div.ifM9O > div > div.kp-header > div.fYOrjf.kp-hc > div.Hhmu2e.wDYxhc.NFQFxe.viOShc.LKPcQc > div > div > h2 > span')
    try:
        name = name.text
    except AttributeError:
        search = driver.find_element_by_css_selector('#tsf > div:nth-child(1) > div.A8SBwf > div.RNNXgb > div > div.a4bIc > input')
        search.clear()
        return False 
    #ドメインと比較
    try:
        respons = rq.get("http://www." + domein)
        url = respons.url
        pre_url = driver.find_element_by_css_selector('#rhs > div > div.kp-blk.knowledge-panel.Wnoohf.OJXvsb > div > div.ifM9O > div > div.kp-header > div.fYOrjf.kp-hc > div.Hhmu2e.wDYxhc.NFQFxe.viOShc.LKPcQc > div > div > div > div:nth-child(1) > a')
        pre_url = pre_url.get_attribute('href')
        if url == pre_url:
            sheet["D" + str(index)].value = name
            fill_yel = pxstyle.PatternFill(patternType='solid', fgColor='FFFF00', bgColor='FFFF00')
            sheet["D" + str(index)].fill = fill_yel
            book.save(file_name)
            print("succsess inport name : " + sheet["D" + str(index)].value)
            search = driver.find_element_by_css_selector('#tsf > div:nth-child(1) > div.A8SBwf > div.RNNXgb > div > div.a4bIc > input')
            search.clear()
    except:
        search = driver.find_element_by_css_selector('#tsf > div:nth-child(1) > div.A8SBwf > div.RNNXgb > div > div.a4bIc > input')
        search.clear()
        return False
"""
    

#会社情報のスクレイピング
driver = webdriver.Chrome(executable_path="./chromedriver_win32/chromedriver.exe")
driver.get("https://www.google.co.jp/maps/@35.3646982,139.5381833,15z?hl=ja")
time.sleep(3)
def com_info(com_name, com_domein, index):
    if sheet["G" + str(index)].value != None:
        return False
    search = driver.find_element_by_xpath('//*[@id="searchboxinput"]')
    search.clear()
    search.send_keys(com_name)
    search = driver.find_element_by_xpath('//*[@id="searchbox-searchbutton"]')
    search.click()
    time.sleep(3)

    #ソースの抽出
    map_data = driver.page_source
    soup = bs(map_data, 'lxml')
    info = soup.find_all(class_="ugiz4pqJLAG__primary-text gm2-body-2")

    #ドメインが合っているか確認
    try:
        domein = info[1].text.strip()
    except:
        return False

    if domein == com_domein:
        pass
    else:
        return False
    #電話
    pattern = r'[\(]{0,1}[0-9]{2,4}[\)\-\(]{0,1}[0-9]{2,4}[\)\-]{0,1}[0-9]{3,4}'
    tel = re.findall(pattern, info[2].text.strip())

    #住所
    address_data = info[0].text.strip()
    all_address = address_data[10:]
    pre_name = re.match('東京都|北海道|(?:京都|大阪)府|.{2,3}県' , all_address)
    address_com = re.split('東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)#県名とそれ以降を分離
    add1 = pre_name.group()#県
    add2 = address_com[1]#それ以降
    
    #指定のセルへ書き込み
    print(tel)
    print(add1)
    print(add2)
    sheet["E" + str(index)].value = tel
    sheet["F" + str(index)].value = add1
    sheet["G" + str(index)].value = add2
    book.save(file_name)


for i in range(307, sheet.max_row):
    company = sheet["D" + str(i)].value
    com_domein = sheet["C" + str(i)].value
    if company != None or company != "取得不可" or company != "不明なエラー":
        try:
            print("writing_data of " + company)    
            write = com_info(company, com_domein, i)
            if write == False:
                print("%s : domein Error." % (sheet["D" + str(i)].value))
        except:
            pass


    
